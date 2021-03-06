#!/usr/bin/env python3
import sys
# funny python's way to add a method to an instance of a class
import types
import thread
import time
import socket
import os
from mido import MidiFile
import random
import threading
import datetime
from openpyxl import load_workbook
from struct import *
def raymap(value, istart, istop, ostart, ostop):
    #wierd = ostart + (ostop - ostart) * (value - istart) / (istop - istart); 
    #print wierd
    return value#wierd

def checkbound(whattype, oidx):
    global ST,AT,TT,BT
    if 67 == oidx:
        oidx = 1
    if 3 == whattype:
        while 0 == BT[oidx]:
            if 66 == oidx:
                oidx = 1
            else:
                oidx += 1
    elif 2 == whattype:
        while 0 == TT[oidx]:
            if 66 == oidx:
                oidx = 1
            else:
                oidx += 1
    elif 1 == whattype:
        while 0 == AT[oidx]:
            if 66 == oidx:
                oidx = 1
            else:
                oidx += 1
    elif 0 == whattype:
        while 0 == ST[oidx]:
            if 66 == oidx:
                oidx = 1
            else:
                oidx += 1
    return oidx                

def raysendto(raystr, raytuple, myport=5005 ):
    global port
    if raytuple == "202":
        port.sendto( raystr, ("192.168.12." + raytuple, myport) )
        if raystr.startswith("M") == False:
            print raystr
    else:
        threading.Timer( 2, port.sendto, [raystr, ("192.168.12." + raytuple, myport)]).start()
    
def rayiii(playidx, mynote=0):
    global thethree
    global rayshift
    iii = 0
    TF = False
    for iii in range(len(thethree)):
        if playidx in thethree[iii]:
            '''
            mypos = mynote-rayshift[playidx-1]+1
            if mypos < 1:
                mypos = 1
            elif mypos > 20:
                mypos = 20
            '''
            if playidx == thethree[iii][0]:
                raysendto("M" + str(iii) + "F", "202", 12345 )
            elif playidx == thethree[iii][1]:
                raysendto("M" + str(iii) + "S", "202", 12345 )
            elif playidx == thethree[iii][2]:
                raysendto("M" + str(iii) + "T", "202", 12345 )
            TF = True
            break
    return TF                
        
def play_midi():
    global isplay
    global myshift
    global boidx,toidx,aoidx,soidx
    global thethree
    global mid
    #workbook = xlsxwriter.Workbook('demo.xlsx')
    #worksheet = workbook.add_worksheet()
    #f = []
    boundary = 0
    #port.flush()
        
    for i in range(1,67):
        raysendto(pack('BB', 249, 3), str(i) )
        time.sleep(0.01)
    time.sleep(4)
    for i in range(34,67):
        raysendto(pack('BB', 225, 0), str(i) )
        raysendto(pack('BB', 225, 0), str(67-i) )
        time.sleep(0.02)
        #f.append(0)
        #worksheet.write(i, 0, 0)
    for message in mid.play():  #Next note from midi in this moment
        isplay = False          #To avoid duplicate doorbell button press during midi play
        if False:
            print(message)
        if 'note_on' == message.type :
            if 0 == message.velocity:
                rayv = pack('BBB', 144, message.note, 0)
                if message.channel == 3:
                    if pickidx[3].has_key(message.note):
                        raysendto(rayv, str(pickidx[3][message.note]) )
                        del pickidx[3][message.note]
                elif message.channel == 2:
                    if pickidx[2].has_key(message.note):
                        raysendto(rayv, str(pickidx[2][message.note]) )
                        del pickidx[2][message.note]
                elif message.channel == 1:
                    if pickidx[1].has_key(message.note):
                        raysendto(rayv, str(pickidx[1][message.note]) )
                        del pickidx[1][message.note]
                elif message.channel == 0:
                    if pickidx[0].has_key(message.note):
                        raysendto(rayv, str(pickidx[0][message.note]) )
                        del pickidx[0][message.note]
                elif message.channel == 11:
                    if pickidx[7].has_key(message.note):
                        raysendto(rayv, str(pickidx[7][message.note]) )
                        del pickidx[7][message.note]
                    if pickidx[11].has_key(message.note):
                        raysendto(rayv, str(pickidx[11][message.note]) )
                        del pickidx[11][message.note]
                elif message.channel == 10:
                    if pickidx[6].has_key(message.note):
                        raysendto(rayv, str(pickidx[6][message.note]) )
                        del pickidx[6][message.note]
                    if pickidx[10].has_key(message.note):
                        raysendto(rayv, str(pickidx[10][message.note]) )
                        del pickidx[10][message.note]
                elif message.channel == 9:
                    if pickidx[5].has_key(message.note):
                        raysendto(rayv, str(pickidx[5][message.note]) )
                        del pickidx[5][message.note]
                    if pickidx[9].has_key(message.note):
                        raysendto(rayv, str(pickidx[9][message.note]) )
                        del pickidx[9][message.note]
                elif message.channel == 8:
                    if pickidx[4].has_key(message.note):
                        raysendto(rayv, str(pickidx[4][message.note]) )
                        del pickidx[4][message.note]
                    if pickidx[8].has_key(message.note):
                        raysendto(rayv, str(pickidx[8][message.note]) )
                        del pickidx[8][message.note]
            else:
                rayv = pack('BBB', 224, message.note, raymap(message.velocity, 0, 127, boundary, 127))
                if message.channel == 3:
                    boidx = checkbound(3,boidx)
                    if True == rayiii(boidx, message.note):
                        raysendto(rayv, str(boidx) )
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(boidx) , str(boidx) )
                    pickidx[3][message.note] = boidx
                    boidx += 1
                elif message.channel == 2:
                    toidx = checkbound(2,toidx)
                    if True == rayiii(toidx, message.note):
                        raysendto(rayv, str(toidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(toidx) , str(toidx))
                    pickidx[2][message.note] = toidx
                    toidx += 1
                elif message.channel == 1:
                    aoidx = checkbound(1,aoidx)
                    if True == rayiii(aoidx, message.note):
                        raysendto(rayv, str(aoidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(aoidx) , str(aoidx))
                    pickidx[1][message.note] = aoidx
                    aoidx += 1
                elif message.channel == 0:
                    soidx = checkbound(0,soidx)
                    if True == rayiii(soidx, message.note):
                        raysendto(rayv, str(soidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(soidx) , str(soidx))
                    pickidx[0][message.note] = soidx
                    soidx += 1
                elif message.channel == 11:
                    boidx = checkbound(3,boidx)
                    if True == rayiii(boidx, message.note):
                        raysendto(rayv, str(boidx) )
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(boidx) , str(boidx) )
                    pickidx[7][message.note] = boidx
                    boidx += 1
                    boidx = checkbound(3,boidx)
                    if True == rayiii(boidx, message.note):
                        raysendto(rayv, str(boidx) )
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(boidx) , str(boidx) )
                    pickidx[11][message.note] = boidx
                    boidx += 1
                elif message.channel == 10:
                    toidx = checkbound(2,toidx)
                    if True == rayiii(toidx, message.note):
                        raysendto(rayv, str(toidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(toidx) , str(toidx))
                    pickidx[6][message.note] = toidx
                    toidx += 1
                    toidx = checkbound(2,toidx)
                    if True == rayiii(toidx, message.note):
                        raysendto(rayv, str(toidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(toidx) , str(toidx))
                    pickidx[10][message.note] = toidx
                    toidx += 1
                elif message.channel == 9:
                    aoidx = checkbound(1,aoidx)
                    if True == rayiii(aoidx, message.note):
                        raysendto(rayv, str(aoidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(aoidx) , str(aoidx))
                    pickidx[5][message.note] = aoidx
                    aoidx += 1
                    aoidx = checkbound(1,aoidx)
                    if True == rayiii(aoidx, message.note):
                        raysendto(rayv, str(aoidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(aoidx) , str(aoidx))
                    pickidx[9][message.note] = aoidx
                    aoidx += 1
                elif message.channel == 8:
                    soidx = checkbound(0,soidx)
                    if True == rayiii(soidx, message.note):
                        raysendto(rayv, str(soidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(soidx) , str(soidx))
                    pickidx[4][message.note] = soidx
                    soidx += 1
                    soidx = checkbound(0,soidx)
                    if True == rayiii(soidx, message.note):
                        raysendto(rayv, str(soidx))
                    else:
                        pass#raysendto("144 " + str(message.note) + " 1 " + str(soidx) , str(soidx))
                    pickidx[8][message.note] = soidx
                    soidx += 1
        elif 'note_off' == message.type :
            rayv = pack('BBB', 128, message.note, message.velocity)
            if message.channel == 3:
                if pickidx[3].has_key(message.note):
                    raysendto(rayv, str(pickidx[3][message.note]) )
                    del pickidx[3][message.note]
            elif message.channel == 2:
                if pickidx[2].has_key(message.note):
                    raysendto(rayv, str(pickidx[2][message.note]) )
                    del pickidx[2][message.note]
            elif message.channel == 1:
                if pickidx[1].has_key(message.note):
                    raysendto(rayv, str(pickidx[1][message.note]) )
                    del pickidx[1][message.note]
            elif message.channel == 0:
                if pickidx[0].has_key(message.note):
                    raysendto(rayv, str(pickidx[0][message.note]) )
                    del pickidx[0][message.note]
            elif message.channel == 11:
                if pickidx[7].has_key(message.note):
                    raysendto(rayv, str(pickidx[7][message.note]) )
                    del pickidx[7][message.note]
                if pickidx[11].has_key(message.note):
                    raysendto(rayv, str(pickidx[11][message.note]) )
                    del pickidx[11][message.note]
            elif message.channel == 10:
                if pickidx[6].has_key(message.note):
                    raysendto(rayv, str(pickidx[6][message.note]) )
                    del pickidx[6][message.note]
                if pickidx[10].has_key(message.note):
                    raysendto(rayv, str(pickidx[10][message.note]) )
                    del pickidx[10][message.note]
            elif message.channel == 9:
                if pickidx[5].has_key(message.note):
                    raysendto(rayv, str(pickidx[5][message.note]) )
                    del pickidx[5][message.note]
                if pickidx[9].has_key(message.note):
                    raysendto(rayv, str(pickidx[9][message.note]) )
                    del pickidx[9][message.note]
            elif message.channel == 8:
                if pickidx[4].has_key(message.note):
                    raysendto(rayv, str(pickidx[4][message.note]) )
                    del pickidx[4][message.note]
                if pickidx[8].has_key(message.note):
                    raysendto(rayv, str(pickidx[8][message.note]) )
                    del pickidx[8][message.note]
    raysendto("EndofGame", "202", 12345 )
    time.sleep(6);
    for i in range(66,33,-1):
        raysendto(pack('BB', 225, 1), str(i) )
        raysendto(pack('BB', 225, 1), str(67-i) )
        time.sleep(0.02)
    for i in range(1,67):
        if 1 == ST[i]:
            raysendto(pack('BBB', 144, 60, 1), str(i) )
        if 1 == AT[i]:
            raysendto(pack('BBB', 144, 48, 1), str(i) )
        if 1 == TT[i]:
            raysendto(pack('BBB', 144, 38, 1), str(i) )
        if 1 == BT[i]:
            raysendto(pack('BBB', 144, 28, 1), str(i) )
        time.sleep(0.02)
    time.sleep(3);
    for i in range(1,67):
        if 1 == ST[i]:
            raysendto(pack('BBB', 144, 60, 0), str(i))
        if 1 == AT[i]:
            raysendto(pack('BBB', 144, 48, 0), str(i))
        if 1 == TT[i]:
            raysendto(pack('BBB', 144, 38, 0), str(i))
        if 1 == BT[i]:
            raysendto(pack('BBB', 144, 28, 0), str(i))
        time.sleep(0.02)
    time.sleep(3);
    for i in range(1,67):
        raysendto(pack('BB', 249, 2), str(i))
        time.sleep(0.01)
    change3(False)
    
def change3(isslider):
    global port
    global howmanyCM
    global thethree
    global set66
    global AmIPlay
    global ws
    global wb
    global myrow
    global mycolumn
    if True == isslider:
        theone = None
        if howmanyCM == 1:
            theone = [2]
        else:
            theone = random.sample(set66, 1)
        AmIPlay = True
        thethree[howmanyCM].append(theone[0]-1)
        thethree[howmanyCM].append(theone[0])
        thethree[howmanyCM].append(theone[0]+1)
        set66.remove(theone[0])
        print set66
        raysendto("CM" + str(howmanyCM) + "F" + str(thethree[howmanyCM][0]) + "S" + str(thethree[howmanyCM][1]) + "T" + str(thethree[howmanyCM][2]), "202", 12345 )
        if 1 == howmanyCM:
            time.sleep(5) # maybe recvfrom another udp string to start it, now it's my control
            thread.start_new_thread(play_midi,())
        howmanyCM += 1
    else:
        for ii in range(len(thethree)):
            if len(thethree[ii]) == 3:
                set66.append(thethree[ii][1])
        for ii in range(len(thethree)):
            for jj in range(len(thethree[ii])):
                thethree[ii].pop()
        print set66
        ws[str(unichr(ord('A') + mycolumn)) + str(myrow)] = howmanyCM-1
        wb.save('/home/oem/Desktop/wechat.xlsx')
        mycolumn = 1
        howmanyCM = 1
        AmIPlay = False
        
run = True

wb = load_workbook('/home/oem/Desktop/wechat.xlsx')
ws = wb.active
myrow = 1
mycolumn = 1
now = datetime.datetime.now()
goout = False
for row in ws.iter_rows('A2:A488'):
    if True == goout:
        break;
    for cell in row:
        myrow = myrow + 1
        if now.year == cell.value.year and now.month == cell.value.month and now.day == cell.value.day: 
            goout = True

data = ""

set66 = [ 2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32, 35, 38, 41, 44, 47, 50, 53, 56, 59, 62, 65 ]

    #27 28 29 30 31 32 33 34 35 36 37 38 39 40 
SS = [0 ,0, 0, 0, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0]
AS = [0 ,0, 0, 1, 1, 1, 0, 0, 0, 1, 1, 0, 0, 0]
TS = [0 ,0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 0]
BS = [1 ,1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1]

           # 1   2   3   4   5   6   7   8   9  10  11  12  13  14  15  16  17  18  19  20
rayshift = [60, 60, 60, 60, 60, 60, 48, 48, 48, 48, 48, 48, 48, 48, 48, 48, 48, 38, 38, 38,
            38, 38, 38, 28, 28, 28, 28, 28, 38, 48, 48, 48, 60, 60, 60, 48, 48, 38, 38, 28, 
            28, 28, 28, 38, 38, 38, 38, 38, 38, 48, 48, 48, 48, 48, 48, 48, 48, 48, 48, 48, 
            60, 60, 60, 60, 60, 60, 60, 48, 38, 60, 60, 48, 38, 28]
            
     #0  1  2  3  4  5  6  7  8  9 10 11 12 13 14 15 16 17 18 19 20 21 22 23 24 25 26 27 28 29 30 31 32 33 34 35 36 37 38 39 40 41 42 43 44 45 46 47 48 49 50 51 52 53 54 55 56 57 58 59 60 61 62 63 64 65 66
ST = [0, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0 ,0 ,0 ,0, 0, 0, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1]
AT = [0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0 ,0 ,0 ,0, 0, 1, 1, 1, 0, 0, 0, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0]
TT = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 0, 0 ,0 ,0 ,0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
BT = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1 ,1 ,1 ,1 ,1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
soidx = 1
aoidx = 1
toidx = 1
boidx = 1
pickidx = [{},{},{},{},{},{},{},{},{},{},{},{}]
slideidx = [{},{},{},{},{},{},{},{},{},{},{},{}]

port = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)      
port.bind(("0.0.0.0", 5005))
port.settimeout(0.005)
mid = None

howmanyCM = 1    
thethree = [[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
tmpthree = [[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
AmIPlay = False
thehr = 17
while run:
    try:
        now = datetime.datetime.now()
        if now.hour > thehr or (now.hour == thehr and now.minute > 38):
            os._exit(1)
        elif now.hour < thehr or (now.hour == thehr and now.minute <= 30):
            data, addr = port.recvfrom(1024)
            print data
            if (data[0:2] == "SS"):
                if False == AmIPlay:
                    if mid is None:
                        midstr = '/home/oem/midi/' + data[2:] + '.mid'
                        if os.path.isfile(midstr):
                            mid = MidiFile(midstr)
                            raysendto("PS" + data[2:], "202", 12345 )
                        else:
                            mid = MidiFile('/home/oem/midi/001.mid')
                    else:
                        raysendto("YSS", "202", 12345 )
                else:
                    raysendto("ES" + data[2:], "202", 12345 )
            elif (data[0:2] == "TM"):
                if False == AmIPlay and mid is None:
                    mid = MidiFile('/home/oem/midi/001.mid')
                if howmanyCM <= 22:
                    if howmanyCM == 1:
                        for row in ws.iter_rows('B1:J1'):
                            for cell in row:
                                if now.hour == cell.value.hour and now.minute == cell.value.minute: 
                                    break
                                mycolumn = mycolumn + 1
                    change3(True)
                else:
                    raysendto("NoCM", "202", 12345 ) #should be something about phone is above three, maybe let webserver to control it
            elif (data[0:2] == "RM"):
                if True == AmIPlay:
                    tmpidx = int(data[2:])
                    for jj in range(len(thethree[tmpidx])):
                        tmpthree[tmpidx].append(thethree[tmpidx][jj])
                        thethree[tmpidx].pop()
            elif (data[0:2] == "BM"):
                if True == AmIPlay:
                    tmpidx = int(data[2:])
                    for jj in range(len(tmpthree[tmpidx])):
                        thethree[tmpidx].append(tmpthree[tmpidx][jj])
                        tmpthree[tmpidx].pop()
                
    except socket.timeout:
        pass                    
    except ValueError:
        pass    
    except IndexError:
        pass
    #each_frame()

server.close()
